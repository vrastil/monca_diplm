import openpyxl
from openpyxl.utils import get_column_letter

def get_keys(sheet):
    """ assumed row '1' and '2' are headers """
    key = None
    all_keys = []
    all_subkeys = []
    for cell in sheet['1']:
        # get key, only for 'Cell'
        # for merged cells, keep last key
        if type(cell).__name__ == 'Cell':
            key = cell.value
        
        # check for subkeys
        subcell = sheet['2'][cell.column - 1]
        subkey = subcell.value
        
        # save key and subkey (None means no subkey)
        all_keys.append(key)
        all_subkeys.append(subkey)

    return all_keys, all_subkeys

    
def get_doc_from_row(row, all_keys, all_subkeys):
    # ignore rows starting with empty cell
    if row[0].value is None:
        return None
    
    # check lenghts
    if len(set([len(row), len(all_keys), len(all_subkeys)])) != 1:
        print("Bad lenghts of field!")
        return None
    
    # go through each cell
    doc = {}
    for cell, key, subkey in zip(row, all_keys, all_subkeys):
        value = cell.value if cell.value is not None else 0
        # no subkeys
        if subkey is None:
            doc[key] = value
        # subkey
        else:
            doc.setdefault(key, {})
            doc[key][subkey] = value
            
    return doc
        
def get_all_data(sheet):
    all_keys, all_subkeys = get_keys(sheet)
    all_docs = []
    
    for row in sheet[3:sheet.max_row-1]:
        doc = get_doc_from_row(row, all_keys, all_subkeys)
        if doc is not None:
            all_docs.append(doc)
            
    return all_docs

def load_xlsx(xlsx_file, sheet_name, coll):
    """ open xlsx file 'xlsx_file', load data from sheets 'sheet_name'
    and saves data to collection 'coll' """
    # load all data
    wb = openpyxl.load_workbook(xlsx_file)
    sheet = wb[sheet_name]
    all_docs = get_all_data(sheet)

     # insert data and get info
    res = coll.insert_many(all_docs)
    print("Vlozeno %i zaznamu." % len(res.inserted_ids))

def set_columns_width(sheet):
    column_widths = []
    for row in sheet.iter_rows(min_row=2):
        for i, cell in enumerate(row):
            length = len(str(cell.value))+2
            if len(column_widths) > i:
                if length > column_widths[i]:
                    column_widths[i] = length
            else:
                column_widths += [length]

    for i, column_width in enumerate(column_widths):
        sheet.column_dimensions[get_column_letter(i+1)].width = column_width

def save_xlsx(xlsx_file, all_data):
    # open file
    wb = openpyxl.Workbook()

    # go throug all data for individual sheets
    for data in all_data:
        sheet = wb.create_sheet(data["sheet_name"])
        if "header" in data:
            sheet.append([data["header"]])
        for data_row in data["data"]:
            sheet.append(data_row)

        # set proper columns width
        set_columns_width(sheet)

    # delete default sheet
    wb.remove_sheet(wb.get_sheet_by_name('Sheet'))

    # save file
    wb.save(xlsx_file)